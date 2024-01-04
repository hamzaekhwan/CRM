from .serializers import *

from rest_framework.views import *
from rest_framework.decorators import *
from rest_framework.response import *
from rest_framework.permissions import *
from django.http import JsonResponse
from CRMapp.models import *
from django.shortcuts import get_object_or_404
from django.utils import timezone
from CRMapp.authentications.permissions import *
from CRMapp.maintenances.serializers import MaintenanceSerializer
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from CRMapp.validators import url_validator
from datetime import datetime
from CRMapp.clients.serializers import ClientSerializer

from openpyxl import Workbook
from django.http import HttpResponse
from CRMapp.models import *
from django.core.exceptions import ObjectDoesNotExist

@api_view(['POST','GET','PUT','DELETE'])
@permission_classes([IsManager | IsManagerMaint | IsEmp])
def contract(request,pk=None):
    if request.method == 'POST' :
        data=request.data

        

        ats=data['ats']
        interest_id=data['interest_id']
        contract_exists = Contract.objects.filter(ats=ats).exists()
        if not contract_exists:

            interest=get_object_or_404(Interest, id=interest_id)


            lift_type=data['lift_type']

            size=data['size']

            floors=data['floors']

            location=data['location']

            signed=data['signed']

            try:
                    url_validator(location)
            except:
                    message = {'detail': 'Invalid URL format for location'}
                    return Response(message, status=status.HTTP_400_BAD_REQUEST)
            new_conrtract=Contract.objects.create(
                                        ats=ats,
                                        interest=interest,
                                        lift_type=lift_type,
                                        size=size,
                                        floors=floors,
                                        location=location,
                                        signed=signed )
            
           
            
            message = {'detail': 'Contract added successfully'}
            return Response(message)
            

        else:
            message = {'detail': 'Contract with this Ats already exists'}
            return Response(message, status=status.HTTP_400_BAD_REQUEST)


    if request.method == 'GET' :
        if pk is not None:
          
            user = get_object_or_404(Contract, id=pk)
            serializer = ContractSerializer(user)
            return Response(serializer.data)
        else:
            query=Contract.objects.all().order_by('-id')

            serializer=ContractSerializer(query,many=True)
            return Response(serializer.data)
    
    if request.method == 'DELETE' :
        
        contract = get_object_or_404(Contract, id=pk)
        contract.delete()
      
        message = {'detail': 'Contract deleted successfully'}
        return Response(message)
    
    if request.method == 'PUT' :
        contract = get_object_or_404(Contract, id=pk)
        data = request.data
        
        contract.ats=data.get('ats', contract.ats)
        contract.lift_type=data.get('lift_type', contract.lift_type)
        contract.size = data.get('size', contract.size)
        contract.floors = data.get('floors', contract.floors)
        location = data.get('location', contract.location)
        contract.signed = data.get('signed', contract.signed)
        if not url_validator(location):
            message = {'detail': 'Invalid URL format for location'}
            return Response(message, status=status.HTTP_400_BAD_REQUEST)

        contract.location = location

        

        contract.save()

        message = {'detail': 'contract updated successfully'}
        return Response(message, status=status.HTTP_200_OK)
    
@api_view(['GET'])  ## pagination api 
@permission_classes([IsManager | IsManagerMaint | IsEmp])
def getcontracts(request):
    query = request.query_params.get('keyword')
    
    if query is None:
        query = ''

    contracts = Contract.objects.filter(ats__icontains=query).order_by('-id')

    # Get the total count before pagination
    total_count = contracts.count()

    page = request.query_params.get('page')
    paginator = Paginator(contracts, 10)

    try:
        contracts = paginator.page(page)
    except PageNotAnInteger:
        contracts = paginator.page(1)
    except EmptyPage:
        contracts = paginator.page(paginator.num_pages)

    if page is None:
        page = 1

    page = int(page)
    
    serializer = ContractSerializer(contracts, many=True)

    # Include the total count in the JSON response
    response_data = {
        'contracts': serializer.data,
        'count': total_count,
        'page': page,
        'pages': paginator.num_pages
    }

    return Response(response_data)


@api_view(['GET'])
@permission_classes([IsManager | IsManagerMaint | IsEmp])
def contract_phases_by_id(request,pk):

    contract = get_object_or_404(Contract, id=pk)

    query=Phase.objects.filter(contract=contract).order_by('-id')
    serializer=PhaseSerializer(query,many=True)
    return JsonResponse(serializer.data,safe=False)

@api_view(['GET'])
@permission_classes([IsManager | IsManagerMaint])
def maintenancelift_contract_by_id(request,pk):

    contract = get_object_or_404(Contract, id=pk)

    query=MaintenanceLift.objects.filter(contract=contract).order_by('-id')
    serializer=ContractMaintenanceSerializer(query,many=True)
    return JsonResponse(serializer.data,safe=False)

@api_view(['GET'])
@permission_classes([IsManager | IsManagerMaint])
def maintenance_contract_by_id(request,pk):

    contract = get_object_or_404(Contract, id=pk)

    query=Maintenance.objects.filter(contract=contract).order_by('-id')
    serializer=MaintenanceSerializer(query,many=True)
    return JsonResponse(serializer.data,safe=False)

@api_view(['GET'])
@permission_classes([IsManager | IsManagerMaint])
def note_contract_by_id(request,pk):

    contract = get_object_or_404(Contract, id=pk)

    query=Note.objects.filter(contract=contract).order_by('-id')
    serializer=NoteSerializer(query,many=True)
    return JsonResponse(serializer.data,safe=False)

@api_view(['GET'])
@permission_classes([IsManager | IsManagerMaint])
def client_info_by_contract_by_id(request,pk):

    contract = get_object_or_404(Contract, id=pk)
    interest=Interest.objects.get(id=contract.interest.id)
    client=interest.client 


    query=Client.objects.get(id=client.id)
    serializer=ClientSerializer(query,many=False)
    return JsonResponse(serializer.data,safe=False)

class DistinctFloorAPIView(APIView):
    def get(self, request, *args, **kwargs):
        distinct_floors = Contract.objects.values_list('floors', flat=True).distinct()
        floor_list = [int(floor) for floor in list(distinct_floors)]
        return Response(floor_list)
    
class DistinctLiftTypeAPIView(APIView):
    def get(self, request, *args, **kwargs):
        distinct_lift_types = Contract.objects.values_list('lift_type', flat=True).distinct()
        lift_type_list = list(distinct_lift_types)
        return Response(lift_type_list)
#api to get contracts by client id
@api_view(['GET'])
@permission_classes([IsManager | IsManagerMaint | IsEmp])
def client(request,pk=None):
   

    client=get_object_or_404(Client,id=pk)
    interests=Interest.objects.filter(client=client)

    
    contract=Contract.objects.filter(interest__in=interests).order_by('-id')
       
        

    serializer=ContractSerializer(contract,many=True)
    return JsonResponse(serializer.data,safe=False)
#################################################################################

@api_view(['POST','PUT','DELETE','GET'])
@permission_classes([IsManager | IsManagerMaint | IsEmp])
def note(request,pk=None):
    if request.method == 'POST' :
        data=request.data
        
        
        contract = get_object_or_404(Contract, id=pk)

        note=data['note']
        date=data['date']
        try:
            attachment=request.FILES.get('attachment')
        except:
            attachment=None

        Note.objects.create(contract=contract,
                                        
                                        note=note,
                                        attachment=attachment,
                                        date=date)
        
        message = {'detail': 'Note added successfully'}
        return Response(message)
    if request.method == 'DELETE' :
        query=get_object_or_404(Note, id=pk)
        query.delete()
        message = {'detail': 'Note deleted successfully'}
        return Response(message) 
    if request.method == 'PUT' :
        note_obj = get_object_or_404(Note, id=pk)
        data = request.data

        
        note_obj.note = data.get('note', note_obj.note)
        note_obj.date = data.get('date', note_obj.date)
        
        
        attachment = request.FILES.get('attachment', None)
        if attachment is not None:
            note_obj.attachment = attachment

        note_obj.save()

        message = {'detail': 'Note updated successfully'}
        return Response(message, status=status.HTTP_200_OK)
  
    if request.method == 'GET' :
        if pk is not None:
            note = get_object_or_404(Note, id=pk)
            serializer = NoteSerializer(note)
            return Response(serializer.data)
        else:
            notes = Note.objects.all().order_by('-id')
            serializer = NoteSerializer(notes, many=True)
            return Response(serializer.data)

@api_view(['GET'])  ## pagination api 
@permission_classes([IsManager | IsManagerMaint | IsEmp])
def getnotes(request):  
    query = request.query_params.get('keyword')
    
    if query is None:
        query = ''

    notes = Note.objects.filter(contract__ats__icontains=query).order_by('-id')

    # Get the total count before pagination
    total_count = notes.count()

    page = request.query_params.get('page')
    paginator = Paginator(notes, 10)

    try:
        notes = paginator.page(page)
    except PageNotAnInteger:
        notes = paginator.page(1)
    except EmptyPage:
        notes = paginator.page(paginator.num_pages)

    if page is None:
        page = 1

    page = int(page)
    
    serializer = NoteSerializer(notes, many=True)

    # Include the total count in the JSON response
    response_data = {
        'notes': serializer.data,
        'count': total_count,
        'page': page,
        'pages': paginator.num_pages
    }

    return Response(response_data)
#################################################################################


@api_view(['POST','PUT','DELETE','GET'])
@permission_classes([IsManager | IsManagerMaint | IsEmp])
def phase(request,pk=None):
    if request.method == 'POST' :
        contract = get_object_or_404(Contract, id=pk)

        data=request.data
        start_date_new_phase=data['start_date'] 
        name=data['name']

        if name not in [choice[1] for choice in PHASES_NAME]:
            message = {'detail': 'this phase name doesnt exists'}
            return Response(message, status=status.HTTP_400_BAD_REQUEST)

        existing_phase = Phase.objects.filter(contract=contract, Name=name).first()
        if existing_phase:
            message = {'detail': 'phase with this name already exists'}
            return Response(message, status=status.HTTP_400_BAD_REQUEST)
        


        Phase.objects.create(
            contract=contract,
            Name=name,
            isActive=True,
            start_date=start_date_new_phase,
        )

        return JsonResponse({'detail': 'phase created successfully'})

    if request.method == 'PUT' :
        phase_obj = get_object_or_404(Phase, id=pk)
        data = request.data

       
        phase_obj.start_date = data.get('start_date', phase_obj.start_date)
        phase_obj.end_date = data.get('end_date', phase_obj.end_date)

        today = timezone.now().date()
        if phase_obj.end_date:
            end_date = datetime.strptime(phase_obj.end_date, "%Y-%m-%d").date()
            if end_date <= today:
                phase_obj.isActive = False
            else:
                phase_obj.isActive = True

        phase_obj.save()

        message = {'detail': 'Phase updated successfully'}
        return Response(message, status=status.HTTP_200_OK)
    
    if request.method == 'DELETE' :
        phase=get_object_or_404(Phase, id=pk)
        phase.delete()
        message = {'detail': 'Phase deleted successfully'}
        return Response(message)

    if request.method =="GET":
        if pk is not None:
            phase = get_object_or_404(Phase, id=pk)
            serializer = PhaseSerializer(phase)
            return Response(serializer.data)
        else:
            phases = Phase.objects.all().order_by('-id')
            serializer = PhaseSerializer(phases, many=True)
            return Response(serializer.data)

@api_view(['GET'])  ## pagination api 
@permission_classes([IsManager | IsManagerMaint | IsEmp])
def getphases(request):
    query = request.query_params.get('keyword')
    
    if query is None:
        query = ''

    phases = Phase.objects.filter(contract__ats__icontains=query).order_by('-id')

    # Get the total count before pagination
    total_count = phases.count()

    page = request.query_params.get('page')
    paginator = Paginator(phases, 10)

    try:
        phases = paginator.page(page)
    except PageNotAnInteger:
        phases = paginator.page(1)
    except EmptyPage:
        phases = paginator.page(paginator.num_pages)

    if page is None:
        page = 1

    page = int(page)
    
    serializer = PhaseSerializer(phases, many=True)

    # Include the total count in the JSON response
    response_data = {
        'phases': serializer.data,
        'count': total_count,
        'page': page,
        'pages': paginator.num_pages
    }

    return Response(response_data)

@api_view(['PUT'])
@permission_classes([IsManager | IsManagerMaint | IsEmp])
def end_phase(request,pk):
    
    Phase.objects.filter(id=pk).update(isActive=False,end_date=timezone.now())
    message = {'detail': 'phase ended successfully'}
    return Response(message)






        
@csrf_exempt  # Use this decorator to exempt CSRF protection for simplicity. Consider enabling CSRF in production.
def export_data_to_excel(request):
    if request.method == 'POST':
        try:
            # Assuming you send the contract IDs as a list in the request body
            contract_ids = request.POST.getlist('contract_ids[]')

            # Retrieve the contracts based on the provided IDs
            queryset = Contract.objects.filter(id__in=contract_ids)

            response = HttpResponse(content_type='application/ms-excel')
            response['Content-Disposition'] = 'attachment; filename="client_data.xlsx"'

            workbook = Workbook()
            worksheet = workbook.active

            # Add column headers for Contract model
            contract_fields = Contract._meta.fields
            contract_headers = [field.verbose_name for field in contract_fields]
            for col_num, header in enumerate(contract_headers, start=1):
                worksheet.cell(row=1, column=col_num, value=header)

            # Add column headers for Interest model
            interest_fields = Interest._meta.fields
            interest_headers = [field.verbose_name for field in interest_fields]
            for col_num, header in enumerate(interest_headers, start=len(contract_headers) + 1):
                worksheet.cell(row=1, column=col_num, value=header)

            # Add column headers for Client model
            client_fields = Client._meta.fields
            client_headers = [field.verbose_name for field in client_fields]
            for col_num, header in enumerate(client_headers, start=len(contract_headers) + len(interest_headers) + 1):
                worksheet.cell(row=1, column=col_num, value=header)

            # Add column headers for MaintenanceLift model
            maintenance_lift_fields = MaintenanceLift._meta.fields
            maintenance_lift_headers = [field.verbose_name for field in maintenance_lift_fields]
            for col_num, header in enumerate(maintenance_lift_headers, start=len(contract_headers) + len(interest_headers) + len(client_headers) + 1):
                worksheet.cell(row=1, column=col_num, value=header)

            

        

            # Add column header for active phase name
            worksheet.cell(row=1, column=len(contract_headers) + 1, value='Active Phase Name')

            # Add data to the worksheet
            for row, contract in enumerate(queryset, start=2):
                # Contract data
                for col_num, field in enumerate(contract_fields, start=1):
                    field_name = field.name
                    cell_value = str(getattr(contract, field_name))
                    worksheet.cell(row=row, column=col_num, value=cell_value)

                # Interest data
                interest_data = contract.interest
                for col_num, field in enumerate(interest_fields, start=len(contract_headers) + 1):
                    field_name = field.name
                    cell_value = str(getattr(interest_data, field_name))
                    worksheet.cell(row=row, column=col_num, value=cell_value)

                # Client data
                client_data = interest_data.client
                for col_num, field in enumerate(client_fields, start=len(contract_headers) + len(interest_headers) + 1):
                    field_name = field.name
                    cell_value = str(getattr(client_data, field_name))
                    worksheet.cell(row=row, column=col_num, value=cell_value)

                # MaintenanceLift data (if available)
                try:
                    maintenance_lift_data = contract.maintenancelift
                    for col_num, field in enumerate(maintenance_lift_fields, start=len(contract_headers) + len(interest_headers) + len(client_headers) + 1):
                        field_name = field.name
                        cell_value = str(getattr(maintenance_lift_data, field_name))
                        worksheet.cell(row=row, column=col_num, value=cell_value)
                except ObjectDoesNotExist:
                    # Handle the case where maintenancelift does not exist
                    pass

                # Get the active phase name for the contract
                active_phase_name = Phase.objects.filter(contract=contract, isActive=True).values('Name').first()
                if active_phase_name:
                    active_phase_name = active_phase_name['Name']
                else:
                    active_phase_name = ''  # Set to empty string if no active phase

                # Add the active phase name to the worksheet
                active_phase_col = len(contract_headers) + 1
                worksheet.cell(row=row, column=active_phase_col, value=active_phase_name)

            workbook.save(response)
            return response

        except ObjectDoesNotExist:
            return JsonResponse({'error': 'One or more contracts do not exist.'}, status=400)

    return JsonResponse({'error': 'Invalid request method.'}, status=400)