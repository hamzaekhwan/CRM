
import csv
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from django.http import HttpResponse
from openpyxl import Workbook
from django.http import HttpResponse
from django.core.files.base import ContentFile
import base64,shortuuid
from CRMapp.models import *
from django.core.exceptions import ObjectDoesNotExist
def export_all_data_to_excel(modeladmin, request, queryset):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="client_data.xlsx"'

    workbook = Workbook()
    worksheet = workbook.active

    # Add column headers for Contract model
    contract_fields = Contract._meta.fields
    contract_headers = [field.verbose_name for field in contract_fields]
    for col_num, header in enumerate(contract_headers, start=1):
        worksheet.cell(row=1, column=col_num, value=header)

    # Add column headers for Interest model
    interest_fields = Interest._meta.fields
    interest_headers = [field.verbose_name for field in interest_fields]
    for col_num, header in enumerate(interest_headers, start=len(contract_headers) + 1):
        worksheet.cell(row=1, column=col_num, value=header)

    # Add column headers for Client model
    client_fields = Client._meta.fields
    client_headers = [field.verbose_name for field in client_fields]
    for col_num, header in enumerate(client_headers, start=len(contract_headers) + len(interest_headers) + 1):
        worksheet.cell(row=1, column=col_num, value=header)

    # Add column headers for MaintenanceLift model
    maintenance_lift_fields = MaintenanceLift._meta.fields
    maintenance_lift_headers = [field.verbose_name for field in maintenance_lift_fields]
    for col_num, header in enumerate(maintenance_lift_headers, start=len(contract_headers) + len(interest_headers) + len(client_headers) + 1):
        worksheet.cell(row=1, column=col_num, value=header)

    

 

    # Add column header for active phase name
    worksheet.cell(row=1, column=len(contract_headers) + 1, value='Active Phase Name')

    # Add data to the worksheet
    for row, contract in enumerate(queryset, start=2):
        # Contract data
        for col_num, field in enumerate(contract_fields, start=1):
            field_name = field.name
            cell_value = str(getattr(contract, field_name))
            worksheet.cell(row=row, column=col_num, value=cell_value)

        # Interest data
        interest_data = contract.interest
        for col_num, field in enumerate(interest_fields, start=len(contract_headers) + 1):
            field_name = field.name
            cell_value = str(getattr(interest_data, field_name))
            worksheet.cell(row=row, column=col_num, value=cell_value)

        # Client data
        client_data = interest_data.client
        for col_num, field in enumerate(client_fields, start=len(contract_headers) + len(interest_headers) + 1):
            field_name = field.name
            cell_value = str(getattr(client_data, field_name))
            worksheet.cell(row=row, column=col_num, value=cell_value)

        # MaintenanceLift data (if available)
        try:
            maintenance_lift_data = contract.maintenancelift
            for col_num, field in enumerate(maintenance_lift_fields, start=len(contract_headers) + len(interest_headers) + len(client_headers) + 1):
                field_name = field.name
                cell_value = str(getattr(maintenance_lift_data, field_name))
                worksheet.cell(row=row, column=col_num, value=cell_value)
        except ObjectDoesNotExist:
            # Handle the case where maintenancelift does not exist
            pass

        # Get the active phase name for the contract
        active_phase_name = Phase.objects.filter(contract=contract, isActive=True).values('Name').first()
        if active_phase_name:
            active_phase_name = active_phase_name['Name']
        else:
            active_phase_name = ''  # Set to empty string if no active phase

        # Add the active phase name to the worksheet
        active_phase_col = len(contract_headers) + 1
        worksheet.cell(row=row, column=active_phase_col, value=active_phase_name)

    workbook.save(response)
    return response


def export_to_excel(modeladmin, request, queryset):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="client_data.xlsx"'

    workbook = Workbook()
    worksheet = workbook.active

    # Get the model fields dynamically
    model = queryset.model
    fields = model._meta.fields

    # Add column headers
    headers = [field.verbose_name for field in fields]
    for col_num, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=col_num, value=header)

    # Add client data
    for row, client in enumerate(queryset, start=2):
        for col_num, field in enumerate(fields, start=1):
            field_name = field.name
            cell_value = str(getattr(client, field_name))
            worksheet.cell(row=row, column=col_num, value=cell_value)

    workbook.save(response)
    return response


def export_to_pdf(modeladmin, request, queryset):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="exported_data.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []


    table_data = []

    # Get the field names
    field_names = [field.name for field in queryset.model._meta.fields]

    # Add column headers
    table_data.append([name.capitalize() for name in field_names])

    # Add client data
    for item in queryset:
        table_data.append([str(getattr(item, field)) for field in field_names])

    # Create the table
    table = Table(table_data, colWidths=60)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    elements.append(table)

    # Build the PDF document
    doc.build(elements)

    return response


def export_to_csv(modeladmin, request, queryset):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="exported_data.csv"'

    # Create a CSV writer
    writer = csv.writer(response)
    
    # Get the field names
    field_names = [field.name for field in queryset.model._meta.fields]

    # Write the header row
    writer.writerow(field_names)

    # Write data rows
    for item in queryset:
        row = [str(getattr(item, field)) for field in field_names]
        writer.writerow(row)

    return response

def convert_base64(code64,name1,name2):    
       
            s = shortuuid.ShortUUID(alphabet="0123456789abcdef")
            otp = s.random(length=5)
            var=code64.split('/')[1]
            image_name =  otp+  '.'+var.split(';')[0]

            extension = image_name.split('.')[1].lower()

            image_name = '{}_{}.{}'.format( name1 , name2, extension)

            imgStr = code64.split(';base64')

            new_image = ContentFile(base64.b64decode(imgStr[1]), name=image_name)

            return new_image